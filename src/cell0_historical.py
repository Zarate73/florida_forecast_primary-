from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURACIÓN
# ============================================================

BASE_DIR = Path.cwd()

if not (BASE_DIR / "Historico").exists():
    BASE_DIR = Path(__file__).resolve().parent if "__file__" in globals() else BASE_DIR


def resolve_election_file(*candidates):
    """Busca un archivo real entre variantes de nombre y rutas relativas."""
    for candidate in candidates:
        path = Path(candidate)
        full_path = path if path.is_absolute() else BASE_DIR / path
        if full_path.exists():
            return full_path

    return BASE_DIR / candidates[0]


FILES = {
    2018: resolve_election_file(
        "Historico/08282018Election.txt",
        "Historico/08282018Election(1).txt",
    ),
    2022: resolve_election_file(
        "Historico/08232022Election.txt",
    ),
}

OUTPUT_FILE = BASE_DIR / "Florida_Governor_Primaries_2018_2022.xlsx"


# ============================================================
# CANDIDATOS QUE QUEREMOS SEPARAR EN COLUMNAS
# ============================================================
#
# Los demás candidatos de cada partido se acumulan en:
#
# Other_Rep_Votes
# Other_Dem_Votes
#
# Esto permite usar exactamente la estructura solicitada para
# 2018 y adaptarla correctamente a 2022.
# ============================================================

CANDIDATES = {

    2018: {
        "REP": {
            "DeSantis": "DeSantis_Votes",
            "Putnam": "Putnam_Votes",
        },

        "DEM": {
            "Gillum": "Gillum_Votes",
            "Graham": "Graham_Votes",
            "Levine": "Levine_Votes",
        },
    },

    2022: {
        "REP": {
            # El script permite que no existan candidatos
            # republicanos GOV en el archivo.
        },

        "DEM": {
            "Crist": "Crist_Votes",
            "Fried": "Fried_Votes",
            "Daniel": "Daniel_Votes",
            "Willis": "Willis_Votes",
        },
    },
}


# ============================================================
# 1. LEER ARCHIVO
# ============================================================

def read_election_file(file_path):
    """
    Lee un TXT electoral separado por tabulaciones.

    Prueba varias codificaciones porque los archivos electorales
    de distintos años pueden venir guardados de manera diferente.
    """

    encodings = [
        "utf-8",
        "cp1252",
        "latin-1",
    ]

    for encoding in encodings:

        try:

            df = pd.read_csv(
                file_path,
                sep="\t",
                encoding=encoding,
                dtype=str,
            )

            print(
                f"OK: {file_path.name} "
                f"[encoding={encoding}] "
                f"[filas={len(df):,}]"
            )

            return df

        except UnicodeDecodeError:
            continue

    raise ValueError(
        f"No fue posible leer el archivo: {file_path}"
    )


# ============================================================
# 2. LIMPIEZA
# ============================================================

def clean_election_data(df):
    """
    Limpia el archivo electoral.

    - Elimina espacios de columnas.
    - Elimina espacios al inicio/final.
    - Convierte CanVotes a entero.
    - Normaliza PartyCode y RaceCode.
    """

    df = df.copy()

    # Limpiar nombres de columnas
    df.columns = df.columns.str.strip()

    # Limpiar columnas de texto
    text_columns = df.select_dtypes(
        include=["object", "str"]  # FIX: evita FutureWarning de pandas 3 (string dtype)
    ).columns

    for column in text_columns:

        df[column] = (
            df[column]
            .fillna("")
            .astype(str)
            .str.strip()
        )

    # Normalizar códigos
    df["RaceCode"] = df["RaceCode"].str.upper()

    df["PartyCode"] = df["PartyCode"].str.upper()

    # Votos a número
    df["CanVotes"] = (
        pd.to_numeric(
            df["CanVotes"],
            errors="coerce"
        )
        .fillna(0)
        .astype(int)
    )

    return df


# ============================================================
# 3. FILTRAR PRIMARIA DE GOBERNADOR
# ============================================================

def filter_governor_primary(df):
    """
    Conserva solamente:

    RaceCode = GOV

    y los partidos:

    REP
    DEM
    """

    governor = df[
        (df["RaceCode"] == "GOV")
        &
        (df["PartyCode"].isin(["REP", "DEM"]))
    ].copy()

    return governor


# ============================================================
# 4. CREAR TABLA DE UN PARTIDO
# ============================================================

def build_party_table(
    governor_df,
    party_code,
    candidate_config,
    counties
):
    """
    Construye resultados por condado para un partido.

    Ejemplo REP 2018:

    County
    DeSantis_Votes
    Putnam_Votes
    Other_Rep_Votes
    Total_Rep_Votes

    Si el partido no aparece en ese año, las columnas
    se generan con valor 0.
    """

    party_df = governor_df[
        governor_df["PartyCode"] == party_code
    ].copy()

    # DataFrame base con todos los condados
    result = pd.DataFrame(
        index=counties
    )

    result.index.name = "County"

    # --------------------------------------------------------
    # Crear columna para cada candidato configurado
    # --------------------------------------------------------

    for last_name, output_column in candidate_config.items():

        candidate_votes = (
            party_df[
                party_df["CanNameLast"]
                .str.casefold()
                == last_name.casefold()
            ]
            .groupby("CountyName")["CanVotes"]
            .sum()
        )

        result[output_column] = (
            candidate_votes
            .reindex(counties)
            .fillna(0)
            .astype(int)
        )

    # --------------------------------------------------------
    # Total de votos del partido
    # --------------------------------------------------------

    total_votes = (
        party_df
        .groupby("CountyName")["CanVotes"]
        .sum()
    )

    total_column = (
        "Total_Rep_Votes"
        if party_code == "REP"
        else "Total_Dem_Votes"
    )

    result[total_column] = (
        total_votes
        .reindex(counties)
        .fillna(0)
        .astype(int)
    )

    # --------------------------------------------------------
    # Otros candidatos
    # --------------------------------------------------------

    candidate_vote_columns = list(
        candidate_config.values()
    )

    if candidate_vote_columns:

        selected_votes = result[
            candidate_vote_columns
        ].sum(axis=1)

    else:

        selected_votes = pd.Series(
            0,
            index=result.index
        )

    other_column = (
        "Other_Rep_Votes"
        if party_code == "REP"
        else "Other_Dem_Votes"
    )

    result[other_column] = (
        result[total_column]
        - selected_votes
    )

    # --------------------------------------------------------
    # Orden
    # --------------------------------------------------------

    result = result[
        candidate_vote_columns
        + [
            other_column,
            total_column,
        ]
    ]

    return result


# ============================================================
# 5. CONSTRUIR TABLA COMPLETA POR AÑO
# ============================================================

def build_year_table(
    governor_df,
    year
):
    """
    Une REP + DEM a nivel de condado.
    """

    # --------------------------------------------------------
    # Obtener todos los condados presentes en GOV
    # --------------------------------------------------------

    counties = sorted(
        governor_df["CountyName"]
        .dropna()
        .loc[
            governor_df["CountyName"] != ""
        ]
        .unique()
    )

    # --------------------------------------------------------
    # Republicanos
    # --------------------------------------------------------

    rep_table = build_party_table(
        governor_df=governor_df,
        party_code="REP",
        candidate_config=CANDIDATES[year]["REP"],
        counties=counties,
    )

    # --------------------------------------------------------
    # Demócratas
    # --------------------------------------------------------

    dem_table = build_party_table(
        governor_df=governor_df,
        party_code="DEM",
        candidate_config=CANDIDATES[year]["DEM"],
        counties=counties,
    )

    # --------------------------------------------------------
    # Unir
    # --------------------------------------------------------

    final = rep_table.join(
        dem_table,
        how="outer",
    )

    final = final.fillna(0)

    # --------------------------------------------------------
    # Total de votos de ambas primarias
    # --------------------------------------------------------

    final["Total_Primary_Votes"] = (
        final["Total_Rep_Votes"]
        + final["Total_Dem_Votes"]
    )

    # --------------------------------------------------------
    # County vuelve a ser columna
    # --------------------------------------------------------

    final = final.reset_index()

    # Agregar año
    final.insert(
        0,
        "Year",
        year
    )

    # --------------------------------------------------------
    # Convertir votos a enteros
    # --------------------------------------------------------

    vote_columns = [
        column
        for column in final.columns
        if column not in [
            "Year",
            "County"
        ]
    ]

    final[vote_columns] = (
        final[vote_columns]
        .astype(int)
    )

    return final


# ============================================================
# 6. RESUMEN ESTATAL
# ============================================================

def build_statewide_summary(
    governor_df,
    year
):
    """
    Calcula el total estatal REAL para todos los candidatos
    encontrados en el archivo.

    No depende de CANDIDATES, por lo que ningún candidato
    queda fuera del resumen estatal.
    """

    df = governor_df.copy()

    # --------------------------------------------------------
    # Construir nombre completo
    # --------------------------------------------------------

    name_columns = [
        "CanNameFirst",
        "CanNameMiddle",
        "CanNameLast",
    ]

    for column in name_columns:

        df[column] = (
            df[column]
            .fillna("")
            .astype(str)
            .str.strip()
        )

    df["Candidate"] = (
        df[name_columns]
        .agg(" ".join, axis=1)
        .str.replace(
            r"\s+",
            " ",
            regex=True
        )
        .str.strip()
    )

    # --------------------------------------------------------
    # Agrupar
    # --------------------------------------------------------

    summary = (
        df
        .groupby(
            [
                "PartyCode",
                "PartyName",
                "Candidate",
            ],
            as_index=False,
        )["CanVotes"]
        .sum()
        .rename(
            columns={
                "CanVotes": "Statewide_Votes"
            }
        )
    )

    # Agregar año
    summary.insert(
        0,
        "Year",
        year
    )

    # --------------------------------------------------------
    # Total del partido para calcular porcentaje
    # --------------------------------------------------------

    summary[
        "Party_Total_Votes"
    ] = (
        summary
        .groupby(
            [
                "Year",
                "PartyCode"
            ]
        )["Statewide_Votes"]
        .transform("sum")
    )

    # --------------------------------------------------------
    # Porcentaje estatal dentro de la primaria
    # --------------------------------------------------------

    summary[
        "Vote_Share"
    ] = (
        summary["Statewide_Votes"]
        / summary["Party_Total_Votes"]
    )

    # --------------------------------------------------------
    # Ranking dentro del partido
    # --------------------------------------------------------

    summary[
        "Party_Rank"
    ] = (
        summary
        .groupby(
            [
                "Year",
                "PartyCode"
            ]
        )["Statewide_Votes"]
        .rank(
            method="dense",
            ascending=False
        )
        .astype(int)
    )

    summary = summary.sort_values(
        [
            "Year",
            "PartyCode",
            "Statewide_Votes",
        ],
        ascending=[
            True,
            True,
            False,
        ]
    )

    return summary.reset_index(drop=True)


# ============================================================
# 7. FORMATO EXCEL
# ============================================================

def format_excel(file_path):
    """
    Aplica formato al Excel usando openpyxl.
    """

    workbook = load_workbook(
        file_path
    )

    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78"
    )

    header_font = Font(
        color="FFFFFF",
        bold=True
    )

    for worksheet in workbook.worksheets:

        # Congelar encabezado
        worksheet.freeze_panes = "A2"

        # Filtro
        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        # ----------------------------------------------------
        # Encabezados
        # ----------------------------------------------------

        for cell in worksheet[1]:

            cell.fill = header_fill

            cell.font = header_font

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        # ----------------------------------------------------
        # Formatos numéricos
        # ----------------------------------------------------

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        for row in range(
            2,
            worksheet.max_row + 1
        ):

            for column_name, column_number in headers.items():

                cell = worksheet.cell(
                    row=row,
                    column=column_number
                )

                if (
                    column_name
                    and "Votes" in str(column_name)
                ):
                    cell.number_format = "#,##0"

                if column_name == "Vote_Share":
                    cell.number_format = "0.00%"

        # ----------------------------------------------------
        # Ancho automático
        # ----------------------------------------------------

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                30
            )

    workbook.save(
        file_path
    )


# ============================================================
# 8. PROGRAMA PRINCIPAL
# ============================================================

def main():

    year_tables = {}

    statewide_summaries = []

    # ========================================================
    # PROCESAR CADA AÑO
    # ========================================================

    for year, file_path in FILES.items():

        print()
        print("=" * 60)
        print(f"PROCESANDO {year}")
        print("=" * 60)

        # ----------------------------------------------------
        # Leer
        # ----------------------------------------------------

        df = read_election_file(
            file_path
        )

        # ----------------------------------------------------
        # Limpiar
        # ----------------------------------------------------

        df = clean_election_data(
            df
        )

        # ----------------------------------------------------
        # GOV solamente
        # ----------------------------------------------------

        governor = filter_governor_primary(
            df
        )

        print(
            f"Filas GOV encontradas: "
            f"{len(governor):,}"
        )

        # ----------------------------------------------------
        # Mostrar partidos encontrados
        # ----------------------------------------------------

        parties_found = sorted(
            governor["PartyCode"]
            .unique()
        )

        print(
            "Partidos GOV encontrados:",
            parties_found
        )

        # ----------------------------------------------------
        # Mostrar candidatos encontrados
        # ----------------------------------------------------

        candidates_found = (
            governor[
                [
                    "PartyCode",
                    "CanNameFirst",
                    "CanNameLast"
                ]
            ]
            .drop_duplicates()
            .sort_values(
                [
                    "PartyCode",
                    "CanNameLast"
                ]
            )
        )

        print(
            "\nCandidatos encontrados:"
        )

        print(
            candidates_found
            .to_string(index=False)
        )

        # ----------------------------------------------------
        # Tabla por condado
        # ----------------------------------------------------

        year_table = build_year_table(
            governor,
            year
        )

        year_tables[year] = year_table

        print(
            f"\nCondados procesados: "
            f"{year_table['County'].nunique()}"
        )

        # ----------------------------------------------------
        # Resumen estatal
        # ----------------------------------------------------

        summary = build_statewide_summary(
            governor,
            year
        )

        statewide_summaries.append(
            summary
        )

    # ========================================================
    # COMBINAR RESÚMENES ESTATALES
    # ========================================================

    statewide_summary = pd.concat(
        statewide_summaries,
        ignore_index=True
    )

    # ========================================================
    # EXPORTAR EXCEL
    # ========================================================

    with pd.ExcelWriter(
        OUTPUT_FILE,
        engine="openpyxl"
    ) as writer:

        # 2018
        year_tables[2018].to_excel(
            writer,
            sheet_name="Governor_Primary_2018",
            index=False,
        )

        # 2022
        year_tables[2022].to_excel(
            writer,
            sheet_name="Governor_Primary_2022",
            index=False,
        )

        # Resumen de ambos años
        statewide_summary.to_excel(
            writer,
            sheet_name="Statewide_Summary",
            index=False,
        )

    # Aplicar formato
    format_excel(
        OUTPUT_FILE
    )

    print()
    print("=" * 60)
    print("PROCESO TERMINADO")
    print("=" * 60)

    print(
        f"\nExcel generado:\n"
        f"{OUTPUT_FILE.resolve()}"
    )


# ============================================================
# EJECUCIÓN
# ============================================================

if __name__ == "__main__":
    main()